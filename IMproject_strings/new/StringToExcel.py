#!/usr/bin/python
import os
from xml.dom.minidom import parse
import xml.dom.minidom
#import xdrlib, xlwt
import sys

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color,Font,Alignment,PatternFill,Border,Side,Protection
from openpyxl.utils import get_column_letter



map_list = []

string_keys = []

stringPath = ''
excel_name = ''

def ParseKey(fileName):
    DOMTree = xml.dom.minidom.parse(fileName)
    collection = DOMTree.documentElement
    if collection:
        values = collection.getElementsByTagName("string")
        
    keys = []
    if values:
        for value in values:
            if value.hasAttribute("name"):
                map_key = value.getAttribute("name")
                keys.append(map_key)

    return keys


def ParseXml(fileName):
    print("begin to parse %s" % (fileName))
    DOMTree = xml.dom.minidom.parse(fileName)
    value_map = {}
    collection = DOMTree.documentElement
	#values = []
    if collection:
        values = collection.getElementsByTagName("string")

    if values:
        for value in values:
            if value.hasAttribute("name"):
                map_key = value.getAttribute("name")

                
                children = value.childNodes;
                if children:
                    map_value = value.childNodes[0].data
                    map_value = map_value.replace('&amp;', '&')
                else:
                    map_value = ""
                
                value_map[map_key] = map_value 
    
    
    return value_map


def write_excel(sheet, keys, value_map, col_index, col_name):
    
    sheet.cell(0, col_index, col_name)
    
    for i in range(0, len(keys)):
        key = list(keys)[i]
        value = value_map[key]
        if not value:
            value = ""
        
        sheet.cell((i+1), col_index, value)
        
        

def moveEnglish2First(list):
	
	if list is None or len(list) <= 0:
		return list

	enFolder = 'values'
	isDel = False
	for i in range(len(list)):
		if list[i] == enFolder:
			del list[i]
			isDel = True
			break

	if isDel:
		list.insert(0, enFolder)

	return list
	

def findRowColByKey(keylist, key):

	if len(key) <= 0 or len(keylist) <= 0:
		return 1

	totalRow = len(keylist)

	row = totalRow

	i = 1
	for val in keylist:
		i = i+1
		if (val is None) or len(val) <= 0:
			row = i
			break
		
		if val == key:
			row = i
			break

	if row == totalRow + 1:
		keylist.append(key)

	return row
	

def isStringTranslated(dirc, inkey):

	if len(dirc) <= 0 or len(inkey) <= 0:
		return False

	for key,value in dirc.items():
		if value is None or len(value) <= 0:
			return False

		for k,v in value.items():
			if k == inkey:
				if v is None or len(v) <= 0:
					return False


	return True
					



def writeStringKeyAndLanguage2Excel(valueFiles, stringPath, excel_name, sheet):

	langs = []
	keys = []

	if len(valueFiles) <= 0 or len(excel_name) <= 0:
		print('write excel get exception##########')
		return langs, keys


	dirc = {}
	for file_name in valueFiles:
		if not file_name.startswith('values'):
			continue

		tmp = stringPath + os.sep + file_name
		if not os.path.isdir(tmp):
			continue

		tmp = tmp + os.sep + 'strings.xml'

		if not os.path.exists(tmp):
			continue

		lang = 'en'    
		ret = file_name.find('-')
		if ret != -1:
			lang = file_name[ret+1:]

		langs.append(lang)

		key_map = ParseKey(tmp)
		if lang == 'en':

			keys = key_map
	
		else:
			dirc[lang] = value_map
	
	tmpKeys = keys

	translateList = []
	noTranslateList = []
	for key in tmpKeys:
		ret = isStringTranslated(dirc, key)
		if ret:
			translateList.append(key)
		else:
			noTranslateList.append(key)
			

	dstkeys = []

	dstkeys.append('Key')

	for k in translateList:
		dstkeys.append(k)

	for k in noTranslateList:
		dstkeys.append(k)

	#write lang
	row = 0
	for key in dstkeys:
		sheet.write(row, 0, key)
		row = row + 1

	'''
	col = 1
	for lang in langs:
		sheet.write(0, col, lang)
		col = col + 1
	'''

	return langs, dstkeys
		

def getTotalLanguages(valueFiles, stringPath):

	nTotal = 0
	if len(valueFiles) <= 0:
		print("value files is null.")
		return nTotal

	for file_name in valueFiles:
	
		if not file_name.startswith('values'):
			continue

		tmp = stringPath + os.sep + file_name
		print(tmp)
		if not os.path.isdir(tmp):
			continue

		tmp = tmp + os.sep + 'strings.xml'

		if not os.path.exists(tmp):
			continue

		nTotal = nTotal + 1

	return nTotal

def getTagChar(src):

	ich = 65
	tag = 'A'

	'''
	if src >= 52 or src < 0:
		print('CHAR is larger than [ 52 ] or smaller than [ 0 ], catch error!')
		sys.exit(0)

	if src < 26:
		ich = 65 + src
		tag = str(chr(ich))
	elif src < 52:
		ich = 65 + src%26
		tag = 'A' + str(chr(ich))

	'''
	tag = get_column_letter(src+1)

	return tag


def getCellTag(row, col):

	tag = getTagChar(col-1)

	dst = tag + str(row)

	return dst


def adaptedExcelCellHeight(excel_name):
	if len(excel_name) <= 0:
		return

	wb = openpyxl.load_workbook(excel_name)
	ws = wb[wb.sheetnames[0]]
	print(ws)

	totalRow = ws.max_row
	totalCol = ws.max_column

	for row in range(2, totalRow+1):
		maxheight = 3
		for c in range(2, totalCol+1):
			i = ws.cell(row, c).value
			columns_length = (len(str(i).encode('utf-8'))  - len(str(i)))/2 + len(str(i))
			#columns_length = len(str(i).encode('utf-8'))
			h = (columns_length / 35 + 1)
			if h > maxheight:
				maxheight = h
			

		ws.row_dimensions[row].bestFit = True
		ws.row_dimensions[row].collapsed = False
		ws.row_dimensions[row].auto_size = True		

		ws.row_dimensions[row].height = maxheight*18


	wb.save(excel_name)
			
	
	



def start():

	# stringPath = os.getcwd() + os.sep + '../palmplay61/src/main' + os.sep + "res"

	stringPath = "/Users/lm188/AndroidStudioProjects/im_android_devlop/reslibrary/src/main/res"

	excel_name = os.getcwd() + os.sep + 'strings2021020201.xlsx'

	try:
		if os.path.exists(excel_name):
			os.remove(excel_name)
	except:
		print("file not exist!")


	valueFiles = os.listdir(stringPath)

	moveEnglish2First(valueFiles)

	#print(valueFiles)

	#total language
	totalLang = getTotalLanguages(valueFiles, stringPath)

	excel_file = openpyxl.Workbook()

	sheet = excel_file.active
	sheet.title = 'Strings'
	sheet.freeze_panes = 'A2'
	sheet.row_dimensions[1].height = 30
	sheet["A1"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
	sheet['A1'].font = openpyxl.styles.Font(name="宋体", size=20, italic=False, color="00CCFF", bold=True)

	for i in range(totalLang+1):
		tag = getTagChar(i)
		sheet.column_dimensions[tag].width = 35

	col_index = 1

	#begin parse end write
	for file_name in valueFiles:
		
		if not file_name.startswith('values'):
			continue

		tmp = stringPath + os.sep + file_name
		if not os.path.isdir(tmp):
			continue

		tmp = tmp + os.sep + 'strings.xml'

		if not os.path.exists(tmp):
			continue

		lang = 'en'    
		ret = file_name.find('-')
		if ret != -1:
			lang = file_name[ret+1:]
		value_map = ParseXml(tmp)
		#map_list.append(value_map)

		if lang == 'en':
			string_keys = ParseKey(tmp)
			sheet.cell(1, 1, 'KEY')
			sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
			sheet['A1'].font = openpyxl.styles.Font(name="Noto Sans CJK SC Regular", size=10, italic=False, color="303030", bold=False)
			fill_heading = PatternFill('solid', fgColor='FFFF00')  # 灰色
			sheet.cell(row=1, column=1).fill = fill_heading


		col_index = col_index + 1


		#write LANG begin
		tag = getCellTag(1, col_index)
		sheet[tag].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
		sheet[tag].font = openpyxl.styles.Font(name="Noto Sans CJK SC Regular", size=14, italic=False, color="303030", bold=True)
		fill_heading = PatternFill('solid', fgColor='BFBFBF')  # 灰色
		fill = PatternFill('solid', fgColor='FF9999')  # 亮粉
		sheet.cell(row=1, column=col_index).fill = fill_heading
		sheet.cell(row=1, column=col_index).fill = fill
		
		#df_len = sheet.apply(lambda x:[(len(str(i).encode('utf-8')) - len(str(i)))/2 + len(str(i)) for i in x], axis=0)
		#df_len_max = df_len.apply(lambda x:max(x),axis=0)


		sheet.cell(1, col_index, lang)
		#write LANG end



		#row = 0
		for key,value in value_map.items():
			#row = row + 1

			tmpRow = findRowColByKey(string_keys, key)

			tag = getCellTag(tmpRow, col_index)
			sheet[tag].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
			sheet[tag].font = openpyxl.styles.Font(name="Noto Sans CJK SC Regular", size=10, italic=False, color="303030", bold=False)

			#key background color begin
			tag2 = getCellTag(tmpRow, 1)
			sheet[tag2].alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")
			sheet[tag2].font = openpyxl.styles.Font(name="Noto Sans CJK SC Regular", size=10, italic=False, color="303030", bold=False)
			fill_heading = PatternFill('solid', fgColor='FFFF00')  # 灰色
			sheet.cell(row=tmpRow, column=1).fill = fill_heading
			#key background color end
			sheet.cell(tmpRow, 1, key)

			sheet.cell(tmpRow, col_index, value)



	#save excel
	excel_file.save(excel_name);


	adaptedExcelCellHeight(excel_name)


	print('\nTotal LANGUAGE: ', col_index-1)
	
	print('\n')

	print("Done.")




if __name__ == '__main__':

	print("Begin Auto gen Excel from string value in res folder....")

	start()

    
    











